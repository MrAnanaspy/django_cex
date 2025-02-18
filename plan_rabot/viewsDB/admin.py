from django.contrib import admin
from .models import Appeal, Detail, Materials, Materials3D, Equipment, AdditionalEexpenses, TimeCosts, Expenses, \
    Assembling, NameMaterials
from django.db.models.signals import post_save, pre_save
from django.dispatch import receiver
from django.utils.html import format_html
import openpyxl
import configparser
import os


class DetailInline(admin.TabularInline):
    model = Assembling

@admin.register(Appeal)
class AppealAdmin(admin.ModelAdmin):
    list_display = ('EAM', 'quantity', 'speed', 'start_time', 'end_time', 'image_tag', 'zayavka_link', 'msk_link')
    list_filter = ('start_time', 'speed', 'machine')
    search_fields = ('EAM__name', 'EAM__EAM', 'start_time')

    def image_tag(self, obj):
            return format_html('<img src="{}" style="max-width:100px; max-height:100px"/>'.format(obj.EAM.photo.url))
    image_tag.short_description = 'Фото/Рендер'
    image_tag.allow_tags = True

    def zayavka_link(self, obj):
        if obj.link and obj.link != '/':
            return format_html('<a href="{}">Ссылка на заявку</a>'.format(obj.link))
        else:
            return "Нет ссылки"

    zayavka_link.allow_tags = True
    zayavka_link.short_description = 'Ссылка'

    def msk_link(self, obj):
        if os.path.isdir('media/{0}/addition'.format(obj.EAM.EAM)):
            os.mkdir('media/{0}/addition'.format(obj.EAM.EAM))
        config = configparser.ConfigParser()  # создаём объекта парсера
        config.read("config.ini")

        wb = openpyxl.load_workbook("media/documents/exel/MSK.xlsx")
        sheet = wb.active
        # № appeal
        sheet.cell(row=10, column=2, value=obj.id)
        # EAM
        sheet.cell(row=10, column=3, value=obj.EAM.EAM)
        # Name
        sheet.cell(row=10, column=7, value=obj.EAM.name)
        # EAM
        sheet.cell(row=12, column=11, value=obj.EAM.size)
        # Quantity
        sheet.cell(row=12, column=12, value=obj.quantity)
        # Material
        if obj.EAM.mater:
            sheet.cell(row=10, column=13, value=obj.EAM.mater.name)
        # № MSK
        sheet.cell(row=6, column=6, value="Маршрутно-сопроводительная карта (МСК) №__" + obj.start_time.strftime('%y') + str(obj.id) + '__')

        dat = []
        # Заготовительня
        if obj.EAM.procurement_work != 0:
            dat.append('Заготовительная')

        if 'electroerosion' in obj.machine:
            dat.append('Электроэрозионная')

        if 'turning_milling' in obj.machine:
            dat.append('Токарно-фрезерная')
        else:
            # Токарная
            if 'turning' in obj.machine:
                dat.append('Токарная')

            # Фрезерная
            if 'milling' in obj.machine:
                dat.append('Фрезерная')


        dat.append('Постобработка')
        dat.append('Упаковка')
        num = 1
        for i in dat:
            sheet.cell(row=15+(num * 2), column=3, value=i)
            sheet.cell(row=15+(num * 2), column=2, value=num)
            num+=1

        wb.save('media/{0}/addition/MSK_{1}.xlsx'.format(obj.EAM.EAM, obj.EAM.EAM))

        return format_html('<a href="{0}">Сгенерировать МСК</a>'.format("/media/" + obj.EAM.EAM + "/addition/MSK_" + obj.EAM.EAM + ".xlsx"))
    zayavka_link.allow_tags = True
    zayavka_link.short_description = 'Ссылка'


@admin.register(Detail)
class DetailAdmin(admin.ModelAdmin):
    list_display = ('EAM', 'name', 'image_tag', 'model_link', 'plan_link')
    list_filter = ('EAM', 'name')
    search_fields = ('EAM', 'name')
    inlines = [DetailInline]

    def image_tag(self, obj):
            return format_html('<img src="{}" style="max-width:100px; max-height:100px"/>'.format(obj.photo.url))
    image_tag.short_description = 'Фото/Рендер'
    image_tag.allow_tags = True

    def model_link(self, obj):
        if obj.model:
            return format_html('<a href="{}" download">Скачать модель</a>'.format(obj.model.url))
        else:
            return "Нет модели"

    model_link.allow_tags = True
    model_link.short_description = 'Модель'


    def plan_link(self, obj):
        if obj.plan:
            return format_html('<a href="{}" download">Скачать чертеж</a>'.format(obj.plan.url))
        else:
            return "Нет чертежа"

    plan_link.allow_tags = True
    plan_link.short_description = 'Чертеж'


'''@admin.register(Materials)
class MaterialAdmin(admin.ModelAdmin):
    list_display = ('name', 'shape', 'width', 'long', 'height', 'certificate', 'melting', 'party')
    list_filter = ('name', 'shape')'''


@admin.register(NameMaterials)
class NameMaterialAdmin(admin.ModelAdmin):
    list_display = ('name',)


@admin.register(Expenses)
class ExpensesAdmin(admin.ModelAdmin):
    list_display = ('time', 'fot', 'tool', 'depreciation', 'electricity')
    list_filter = ('time',)


'''@admin.register(Materials3D)
class Material3dAdmin(admin.ModelAdmin):
    list_display = ('name', 'printing_technology', 'quantity')
    list_filter = ('name', 'printing_technology')'''


#AdditionalEexpenses
'''@admin.register(AdditionalEexpenses)
class AdditionalEexpenseAdmin(admin.ModelAdmin):
    list_display = ('designation', 'name_equip', 'name_detail', 'img_detail')

    def name_equip(self, obj):
        if Equipment.objects.filter(designation=obj.designation).exists():
            result = Equipment.objects.get(designation=obj.designation).name
        else:
            result = '-'
        return result

    name_equip.allow_tags = True
    name_equip.short_description = 'Название инструмента'

    def name_detail(self, obj):
        if Appeal.objects.filter(id=obj.appeal_id).exists():
            eam = Appeal.objects.get(id=obj.appeal_id).EAM
            result = Detail.objects.get(EAM=eam).name
        else:
            result = '-'
        return result

    name_detail.allow_tags = True
    name_detail.short_description = 'Название детали'

    def img_detail(self, obj):
        if Appeal.objects.filter(id=obj.appeal_id).exists():
            eam = Appeal.objects.get(id=obj.appeal_id).EAM
            return format_html('<img src="{}" style="max-width:100px; max-height:100px"/>'
                               .format(Detail.objects.get(EAM=eam).photo.url))

    img_detail.allow_tags = True
    img_detail.short_description = 'Фото/Рендер детали'
'''

#Equipment
'''@admin.register(Equipment)
class EquipmentAdmin(admin.ModelAdmin):
    list_display = ('name', 'designation', 'type', 'quantity')
    list_filter = ('name', 'type')'''


'''@admin.register(TimeCosts)
class TimeCostsAdmin(admin.ModelAdmin):
    list_display = ('name_detail', 'img_detail', 'twt', 'mwt', 'tmwt', 'procurement_work')

    def name_detail(self, obj):
        if Appeal.objects.filter(id=obj.appeal_id_id).exists():
            result = Appeal.objects.get(id=obj.appeal_id_id).EAM.name
        else:
            result = '-'
        return result

    def img_detail(self, obj):
        if Appeal.objects.filter(id=obj.appeal_id_id).exists():
            return format_html('<img src="{}" style="max-width:100px; max-height:100px"/>'
                               .format(Appeal.objects.get(id=obj.appeal_id_id).EAM.photo.url))

    img_detail.allow_tags = True
    img_detail.short_description = 'Фото/Рендер детали' '''


@receiver(post_save, sender=Detail)
def generate_data(sender, instance, created, **kwargs):
    if created:
        if not instance.photo:
            instance.photo='/no_photo.png'
        instance.save()


@receiver(post_save, sender=AdditionalEexpenses)
def delete_equip(sender, instance, created, **kwargs):
    if created:
        num = Equipment.objects.get(designation=instance.designation).quantity
        Equipment.objects.filter(designation=instance.designation).update(quantity=(num - 1))


@receiver(post_save, sender=Materials)
def add_material(sender, instance, created, **kwargs):
    if created:
        pass

# Генерация TimeCost object при создании Apeal
@receiver(post_save, sender=Appeal)
def add_appeal(sender, instance, created, **kwargs):
    if created:
        '''TimeCosts.objects.create(
                                appeal_id_id = instance.id,
                                 twt= instance.EAM.twt,
                                 twd=instance.EAM.twd,
                                 mwt=instance.EAM.mwt,
                                 mwd=instance.EAM.mwd,
                                 tmwt=instance.EAM.tmwt,
                                 tmwd=instance.EAM.tmwd,
                                 procurement_work=instance.EAM.procurement_work
                                 )'''
        os.mkdir('media/{0}/addition'.format(instance.EAM.EAM))