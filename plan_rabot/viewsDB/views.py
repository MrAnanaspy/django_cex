import openpyxl
from django.shortcuts import render
from openpyxl.styles import PatternFill
import configparser
from .models import Appeal, TimeCosts, Expenses
import datetime


def get_data(request):
    generate_production_plan()
    data = "media/documents/exel/stata_done.xlsx"
    return render(request, "stata.html", context={'exel':data}) #, context={'data':data}



def generate_production_plan():
    config = configparser.ConfigParser()  # создаём объекта парсера
    config.read("config.ini")

    # Load the workbook
    wb = openpyxl.load_workbook("media/documents/exel/stata.xlsx")

    # Select the active sheet
    sheet = wb.active

    count = 1
    # Read and print the data
    for appeal in Appeal.objects.all():

        #color
        color = 'FFFFFF'
        if appeal.ready_status and appeal.material and appeal.material.status != 'in_stock':
            color = 'FFFF00'
        elif appeal.end_time and datetime.date.today() < appeal.end_time:
            color = '00B0F0'
        else:
            color = '00D050'

        #objects
        time_cost = TimeCosts.objects.get(appeal_id_id=appeal.id)
        if appeal.start_time and Expenses.objects.filter(time__year=appeal.start_time.year, time__month=appeal.start_time.month).exists():
            expenses = Expenses.objects.get(time__year=appeal.start_time.year, time__month=appeal.start_time.month)
        else:
            expenses = Expenses()


        #const
        time = time_cost.twt + time_cost.twd + time_cost.mwt + time_cost.mwd + time_cost.ewt + time_cost.ewd + time_cost.procurement_work
        time_a = time_cost.twt + time_cost.mwt + time_cost.ewt
        time_b = time_cost.twt + time_cost.mwt + time_cost.tmwt + time_cost.twd + time_cost.mwd + time_cost.ewd
        count += 1

        # add id
        sheet.cell(row=count, column=1, value=appeal.id)
        sheet.cell(row=count, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # add EAM
        sheet.cell(row=count, column=2, value=appeal.EAM.EAM)
        sheet.cell(row=count, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # add name
        sheet.cell(row=count, column=3, value=appeal.EAM.name)
        sheet.cell(row=count, column=3).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # add quantity
        sheet.cell(row=count, column=4, value=appeal.quantity)
        sheet.cell(row=count, column=4).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # add time_work
        sheet.cell(row=count, column=5, value=time)
        sheet.cell(row=count, column=5).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # add material_price
        sheet.cell(row=count, column=6, value=appeal.material_price)
        sheet.cell(row=count, column=6).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # косвенные затраты
        production_time = 0
        production_time_a = 0
        eq = 0
        tool_prise = expenses.tool
        if time_a != 0 and tool_prise != 0 and expenses.time and Appeal.objects.filter(start_time__year=expenses.time.year, start_time__month=expenses.time.month).exists():
            for ap in Appeal.objects.filter(start_time__year=expenses.time.year, start_time__month=expenses.time.month):
                t_c = TimeCosts.objects.get(appeal_id=ap.id)
                production_time += (t_c.twt + t_c.mwt + t_c.tmwt) * ap.quantity
                production_time_a += (t_c.twt + t_c.mwt + t_c.tmwt + t_c.twd + t_c.mwd + t_c.tmwd + t_c.procurement_work) * ap.quantity
            eq = round((time_a * appeal.quantity / production_time * tool_prise), 3)
        eq += appeal.equipment_price
        sheet.cell(row=count, column=7, value=eq)
        sheet.cell(row=count, column=7).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        #  затраты на персонал
        pers_cost = 0.0
        if time_a !=0 and expenses.fot and production_time_a !=0:
            pers_cost = round(expenses.fot * (time_b * appeal.quantity / production_time_a), 3)
        sheet.cell(row=count, column=8, value=pers_cost)
        sheet.cell(row=count, column=8).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        #  затраты на амортизацию
        amort = 0.0
        if time_a != 0 and production_time != 0:
            amort = round((time_a * appeal.quantity / production_time) * expenses.depreciation, 3)
        sheet.cell(row=count, column=9, value=amort)
        sheet.cell(row=count, column=9).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        #  затраты на электроэнергию
        electricity = 0
        if (TimeCosts.objects.filter(appeal_id=appeal.id).exists()
                and expenses.electricity
                and time != 0
                and production_time != 0):
            electricity = round(expenses.electricity * (time_a * appeal.quantity / production_time), 3)
        sheet.cell(row=count, column=10, value=electricity)
        sheet.cell(row=count, column=10).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Себестоимость
        cost_price = round((pers_cost + electricity) / appeal.quantity + appeal.material_price, 3)
        sheet.cell(row=count, column=11, value=cost_price)
        sheet.cell(row=count, column=11).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # start_time
        if appeal.start_time is not None:
            sheet.cell(row=count, column=12, value=appeal.start_time)
        else:
            sheet.cell(row=count, column=12, value=datetime.date.today())
        sheet.cell(row=count, column=12).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # end_time
        if appeal.end_time:
            sheet.cell(row=count, column=13, value=appeal.end_time)
        else:
            sheet.cell(row=count, column=13, value=datetime.date.today())
        sheet.cell(row=count, column=13).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save("media/documents/exel/stata_done.xlsx")

